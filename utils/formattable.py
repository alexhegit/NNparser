# -*- coding: utf-8 -*-
"""
Created on Tue Jul 14 18:10:10 2020

@author: sts
"""

import  pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def SumTable(paraout):
    workbook = load_workbook(filename=paraout)
    sheet = workbook['details']#.active 
    if sheet['A1'].value==None:
        sheet.delete_rows(idx=1)
        sheet.delete_cols(idx=1)
    data=sheet.values
    columns = next(data)[0:]
    df=pd.DataFrame(data,columns=columns)
    # df = df.reset_index()
    
    
    maxVal=[['']*3,['']*3]
   
    # max size
    maxVal[0][0] = df['SizeO'].fillna(0).astype('int64').max() # activation
    maxVal[0][1] = df['SizeW'].fillna(0).astype('int64').max()
    # max ops 
    maxVal[1][0] = df['OpGemm'].fillna(0).astype('int64').max()
    
    sumlst = [
                ['Total Activations(MB):',df['SizeO'].fillna(0).astype('int64').sum()/(1000**2)],
                ['Total Weights(MB):',df['SizeW'].fillna(0).astype('int64').sum()/(1000**2)],
                ['Total Gemm (G_ops):',df['OpGemm'].fillna(0).astype('int64').sum()/(1000**3)]
              ]
    
    if "Summary" in set(workbook.sheetnames):
        sheet = workbook["Summary"]
    else:
        sheet = workbook.create_sheet("Summary")
    for i in range(1,4,1):
        sheet["A{}".format(i)] = sumlst[i-1][0]
        sheet["B{}".format(i)] = sumlst[i-1][1]
        
    sheet.insert_rows(idx=1) 
    summarystr = 'Model Statistics:'
    sheet['A1']= summarystr
    ft= Font(b=True)
    sheet['A1'].font=ft
    sheet.column_dimensions['A'].width = max(25,len(summarystr))
    workbook.save(paraout)
    
    return maxVal

    
def FormatTable(tablename,maxValues):
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule
    maxSiActi = maxValues[0][0]
    maxSiWeig = maxValues[0][1]
    maxOpGemm = maxValues[1][0]
    
    workbook = load_workbook(filename=tablename)
    sheet = workbook['details']#.active
    
    if sheet['A1'].value==None:
        sheet.delete_rows(idx=1)
        sheet.delete_cols(idx=1)
    sheet.freeze_panes = "C2"
    
    # row 0: Grey bkcolor, Bold font
    fil = PatternFill("solid", fgColor="00C0C0C0")
    ft= Font(b=True)
    for cell in list(sheet.rows)[0]:
        cell.fill = fil
        cell.font = ft
        if cell.value=='SizeO':
            so=cell.column_letter
        if cell.value=='SizeW':
            sw=cell.column_letter
        if cell.value=='OpGemm':
            og=cell.column_letter
            
    # Max activation row with red
    background = PatternFill(bgColor="00FF0000")
    myrule= CellIsRule(operator='equal', formula=['{}'.format(maxSiActi)], stopIfTrue=True, fill=background)
    sheet.conditional_formatting.add(so+'{}:'.format(sheet.min_row)+so+'{}'.format(sheet.max_row), myrule)
    
        # Max activation row with pink
    background = PatternFill(bgColor="00FF00FF")
    myrule= CellIsRule(operator='equal', formula=['{}'.format(maxSiWeig)], stopIfTrue=True, fill=background)
    sheet.conditional_formatting.add(sw+'{}:'.format(sheet.min_row)+sw+'{}'.format(sheet.max_row), myrule)
    
    #  Max Ops Gemm row with green
    background = PatternFill(bgColor="0000FF00")
    myrule= CellIsRule(operator='equal', formula=['{}'.format(maxOpGemm)], stopIfTrue=True, fill=background)
    sheet.conditional_formatting.add(og+'{}:'.format(sheet.min_row)+og+'{}'.format(sheet.max_row), myrule)
        
    
    workbook.save(tablename)